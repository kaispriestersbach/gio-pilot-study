"""
AP5 Hilfsskript: Simuliere Annotationen fuer Pipeline-Test.

Fuellt das Annotations-Spreadsheet mit simulierten Experten-Annotationen,
so dass AP5 (Auswertung) getestet werden kann.

Die simulierten Annotationen haben:
  - ~80% Uebereinstimmung bei retrieval_judgment
  - ~60% Uebereinstimmung bei gio_mode
  - ~70% Uebereinstimmung bei gn_level

Output: output/annotation_spreadsheet.xlsx (ueberschrieben mit simulierten Daten)
"""

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config


def simulate_rater(n_prompts, seed, baseline_predictions=None):
    """Simuliere Annotationen eines Raters."""
    rng = np.random.default_rng(seed)

    modes = config.DROPDOWN_GIO_MODES  # ["1.1", "1.2", ..., "3.2"]
    gn_levels = config.DROPDOWN_GN_LEVELS  # ["Low", "Medium", "High"]
    retrievals = config.DROPDOWN_RETRIEVAL  # ["Yes", "No"]
    confidences = config.DROPDOWN_CONFIDENCE  # ["1", ..., "5"]

    annotations = {
        "gio_mode": [],
        "i_gap": [],
        "t_decay": [],
        "e_spec": [],
        "v_volatility": [],
        "gn_level": [],
        "retrieval_judgment": [],
        "confidence": [],
    }

    for i in range(n_prompts):
        annotations["gio_mode"].append(rng.choice(modes))
        annotations["i_gap"].append(rng.choice(gn_levels))
        annotations["t_decay"].append(rng.choice(gn_levels))
        annotations["e_spec"].append(rng.choice(gn_levels))
        annotations["v_volatility"].append(rng.choice(gn_levels))
        annotations["gn_level"].append(rng.choice(gn_levels))
        annotations["retrieval_judgment"].append(rng.choice(retrievals))
        annotations["confidence"].append(rng.choice(confidences))

    return annotations


def simulate_correlated_raters(n_prompts, agreement_rate=0.75, seed_a=42, seed_b=123):
    """Simuliere zwei Rater mit einer bestimmten Uebereinstimmungsrate."""
    rng = np.random.default_rng(seed_a)

    modes = config.DROPDOWN_GIO_MODES
    gn_levels = config.DROPDOWN_GN_LEVELS
    retrievals = config.DROPDOWN_RETRIEVAL
    confidences = config.DROPDOWN_CONFIDENCE

    ann_a = {k: [] for k in ["gio_mode", "i_gap", "t_decay", "e_spec",
                               "v_volatility", "gn_level", "retrieval_judgment", "confidence"]}
    ann_b = {k: [] for k in ann_a}

    for i in range(n_prompts):
        # Rater A waehlt
        mode_a = rng.choice(modes)
        gn_a = rng.choice(gn_levels)
        ret_a = rng.choice(retrievals)
        conf_a = rng.choice(confidences)
        igap_a = rng.choice(gn_levels)
        tdecay_a = rng.choice(gn_levels)
        espec_a = rng.choice(gn_levels)
        vvol_a = rng.choice(gn_levels)

        # Rater B: mit agreement_rate Wahrscheinlichkeit gleich, sonst zufaellig
        if rng.random() < agreement_rate:
            mode_b = mode_a
        else:
            mode_b = rng.choice(modes)

        if rng.random() < agreement_rate:
            gn_b = gn_a
        else:
            gn_b = rng.choice(gn_levels)

        if rng.random() < 0.8:  # Retrieval hat hoehere Uebereinstimmung
            ret_b = ret_a
        else:
            ret_b = rng.choice(retrievals)

        conf_b = rng.choice(confidences)

        if rng.random() < 0.7:
            igap_b = igap_a
        else:
            igap_b = rng.choice(gn_levels)

        if rng.random() < 0.7:
            tdecay_b = tdecay_a
        else:
            tdecay_b = rng.choice(gn_levels)

        if rng.random() < 0.7:
            espec_b = espec_a
        else:
            espec_b = rng.choice(gn_levels)

        if rng.random() < 0.7:
            vvol_b = vvol_a
        else:
            vvol_b = rng.choice(gn_levels)

        for ann, vals in [(ann_a, [mode_a, igap_a, tdecay_a, espec_a, vvol_a, gn_a, ret_a, conf_a]),
                          (ann_b, [mode_b, igap_b, tdecay_b, espec_b, vvol_b, gn_b, ret_b, conf_b])]:
            for key, val in zip(ann.keys(), vals):
                ann[key].append(val)

    return ann_a, ann_b


def main():
    print("=" * 60)
    print("AP5: Simulierte Annotationen eintragen (Pipeline-Test)")
    print("=" * 60)

    xlsx_path = config.ANNOTATION_XLSX_PATH
    if not xlsx_path.exists():
        print(f"FEHLER: {xlsx_path} nicht gefunden. Bitte zuerst AP4 ausfuehren.")
        sys.exit(1)

    wb = load_workbook(xlsx_path)

    # Rater-Sheets lesen
    for sheet_name in ["Rater_A", "Rater_B"]:
        ws = wb[sheet_name]
        # Header-Zeile finden
        headers = [cell.value for cell in ws[1]]
        print(f"\n  {sheet_name}: {ws.max_row - 1} Prompts, Spalten: {headers}")

    # Anzahl Prompts (ohne Header)
    n_prompts = wb["Rater_A"].max_row - 1

    # Simulierte korrelierte Annotationen generieren
    print(f"\n  Generiere simulierte Annotationen fuer {n_prompts} Prompts...")
    ann_a, ann_b = simulate_correlated_raters(n_prompts, agreement_rate=0.70)

    # In Excel eintragen
    annotation_cols = ["gio_mode", "i_gap", "t_decay", "e_spec",
                       "v_volatility", "gn_level", "retrieval_judgment", "confidence"]

    for sheet_name, annotations in [("Rater_A", ann_a), ("Rater_B", ann_b)]:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        for col_name in annotation_cols:
            if col_name not in headers:
                print(f"  WARNUNG: Spalte '{col_name}' nicht in {sheet_name} gefunden!")
                continue
            col_idx = headers.index(col_name) + 1  # 1-indexed

            for row_idx, value in enumerate(annotations[col_name], start=2):  # Skip header
                ws.cell(row=row_idx, column=col_idx, value=value)

        print(f"  {sheet_name}: {len(annotation_cols)} Spalten ausgefuellt.")

    # Speichern
    wb.save(xlsx_path)
    print(f"\n  Gespeichert: {xlsx_path}")
    print("  Naechster Schritt: make ap5")


if __name__ == "__main__":
    main()
