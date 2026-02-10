"""
AP4: Annotations-Spreadsheet erstellen.

Erstellt eine Excel-Datei mit folgenden Sheets:
  - Codebook: GIO-Mode-Definitionen und GN-Variablen
  - Kalibrierung: 5 Kalibrierungs-Prompts
  - Rater_A / Rater_B: Annotations-Sheets mit Dropdowns (unterschiedliche Reihenfolge)
  - Baseline: Keyword-Baseline-Vorhersagen
  - Metadaten: Herkunft und Block-Zuordnung (nur fuer Studienleitung)

Input:  data/sampled_prompts.csv, data/baseline_predictions.csv
Output: output/annotation_spreadsheet.xlsx
"""

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config

# -------------------------------------------------------------------------
# Styling-Konstanten
# -------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
CALIBRATION_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
META_TAB_COLOR = "FF0000"
ASKING_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DOING_FILL = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
ACTING_FILL = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
SECTION_FONT = Font(bold=True, size=12)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_codebook_sheet(wb):
    """Erstelle das Codebook-Sheet mit GIO-Definitionen."""
    ws = wb.create_sheet("Codebook")
    ws.sheet_properties.tabColor = "4472C4"

    row = 1

    # Titel
    ws.cell(row=row, column=1, value="GIO Pilot Study — Codebook").font = Font(bold=True, size=14)
    row += 2

    # ----- GIO-Mode-Definitionen -----
    ws.cell(row=row, column=1, value="GIO-Mode-Definitionen (Table 1)").font = SECTION_FONT
    row += 1

    # Header
    headers = ["Mode", "Name", "Kategorie", "GN-Level", "Definition",
               "Beispiel 1", "Beispiel 2", "Abgrenzungsregel"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGNMENT
    row += 1

    # Daten
    category_fills = {"ASKING": ASKING_FILL, "DOING": DOING_FILL, "ACTING": ACTING_FILL}
    for mode_id in config.DROPDOWN_GIO_MODES:
        mode = config.GIO_MODES[mode_id]
        fill = category_fills.get(mode["category"], None)
        values = [
            mode_id,
            mode["name"],
            mode["category"],
            mode["gn_level"],
            mode["definition"],
            mode["examples"][0],
            mode["examples"][1],
            mode["differentiation"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = WRAP_ALIGNMENT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
        row += 1

    row += 2

    # ----- GN-Variablen -----
    ws.cell(row=row, column=1, value="GN-Variablen-Definitionen (Table 4)").font = SECTION_FONT
    row += 1

    headers_gn = ["Variable", "Name", "Definition",
                  "Anker: Low", "Anker: Medium", "Anker: High"]
    for col, h in enumerate(headers_gn, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGNMENT
    row += 1

    for var_key, var in config.GN_VARIABLES.items():
        values = [
            var_key,
            var["name"],
            var["definition"],
            var["anchors"]["Low"],
            var["anchors"]["Medium"],
            var["anchors"]["High"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = WRAP_ALIGNMENT
            cell.border = THIN_BORDER
        row += 1

    row += 2

    # Sonderregel Implicit Demand (Platzhalter)
    ws.cell(row=row, column=1,
            value="Sonderregel: Implicit Demand (Scenario C)").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="[PLATZHALTER — Wird von Kai finalisiert]").font = Font(italic=True, color="FF0000")
    ws.cell(row=row + 1, column=1,
            value="Prompts, die oberflaechlich keine Fakten verlangen, aber implizit aktuelle "
                  "Informationen brauchen. Kein Fragewort, keine Signalwoerter wie 'aktuell' "
                  "oder 'Daten', aber eine valide Antwort erfordert trotzdem externes Wissen.")

    # Spaltenbreiten
    col_widths = [8, 22, 12, 12, 45, 45, 45, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_rater_sheet(wb, sheet_name, study_prompts, seed):
    """Erstelle ein Rater-Sheet mit randomisierter Reihenfolge und Dropdowns."""
    ws = wb.create_sheet(sheet_name)

    # Prompt-Reihenfolge randomisieren
    rng = np.random.default_rng(seed)
    shuffled = study_prompts.sample(frac=1, random_state=int(rng.integers(0, 2**31))).reset_index(drop=True)

    # Header
    headers = [
        ("prompt_id", 8),
        ("prompt_text", 55),
        ("gio_mode", 12),
        ("i_gap", 10),
        ("t_decay", 10),
        ("e_spec", 10),
        ("v_volatility", 12),
        ("gn_level", 10),
        ("retrieval_judgment", 18),
        ("confidence", 12),
        ("notes", 40),
    ]

    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    # Daten einfuellen
    for i, (_, row_data) in enumerate(shuffled.iterrows(), start=2):
        # prompt_id (gesperrt)
        cell = ws.cell(row=i, column=1, value=row_data["prompt_id"])
        cell.fill = LOCKED_FILL
        cell.protection = Protection(locked=True)

        # prompt_text (gesperrt)
        cell = ws.cell(row=i, column=2, value=row_data["prompt_text"])
        cell.fill = LOCKED_FILL
        cell.protection = Protection(locked=True)
        cell.alignment = WRAP_ALIGNMENT

        # Restliche Spalten leer lassen (fuer Rater)
        for col in range(3, 12):
            ws.cell(row=i, column=col).protection = Protection(locked=False)

    n_rows = len(shuffled)

    # Dropdowns (DataValidation)
    dv_mode = DataValidation(
        type="list",
        formula1=f'"{",".join(config.DROPDOWN_GIO_MODES)}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Ungueltiger Modus",
        error="Bitte einen GIO-Modus aus der Liste waehlen.",
    )
    dv_mode.sqref = f"C2:C{n_rows + 1}"
    ws.add_data_validation(dv_mode)

    dv_gn = DataValidation(
        type="list",
        formula1=f'"{",".join(config.DROPDOWN_GN_LEVELS)}"',
        allow_blank=True,
    )
    # i_gap, t_decay, e_spec, v_volatility, gn_level (Spalten D-H)
    dv_gn.sqref = f"D2:H{n_rows + 1}"
    ws.add_data_validation(dv_gn)

    dv_retrieval = DataValidation(
        type="list",
        formula1=f'"{",".join(config.DROPDOWN_RETRIEVAL)}"',
        allow_blank=True,
    )
    dv_retrieval.sqref = f"I2:I{n_rows + 1}"
    ws.add_data_validation(dv_retrieval)

    dv_confidence = DataValidation(
        type="list",
        formula1=f'"{",".join(config.DROPDOWN_CONFIDENCE)}"',
        allow_blank=True,
    )
    dv_confidence.sqref = f"J2:J{n_rows + 1}"
    ws.add_data_validation(dv_confidence)

    # Sheet Protection: nur Annotationsspalten editierbar
    ws.protection.sheet = True
    ws.protection.password = ""  # Kein Passwort, nur visueller Schutz

    # Freeze Panes (Header + Prompt-Text sichtbar beim Scrollen)
    ws.freeze_panes = "C2"

    return ws


def create_calibration_sheet(wb, calib_prompts):
    """Erstelle das Kalibrierungs-Sheet."""
    ws = wb.create_sheet("Kalibrierung")
    ws.sheet_properties.tabColor = "FFC000"

    # Hinweis
    ws.cell(row=1, column=1,
            value="KALIBRIERUNG — Diese Prompts fliessen NICHT in die Auswertung ein.").font = Font(
        bold=True, size=12, color="996600")
    ws.cell(row=2, column=1,
            value="Bitte gemeinsam im Kalibrierungstreffen besprechen.").font = Font(italic=True)

    row = 4

    # Header fuer Rater A
    ws.cell(row=row, column=1, value="prompt_id").font = HEADER_FONT
    ws.cell(row=row, column=2, value="prompt_text").font = HEADER_FONT
    ws.cell(row=row, column=3, value="Rater A: gio_mode").font = HEADER_FONT
    ws.cell(row=row, column=4, value="Rater A: gn_level").font = HEADER_FONT
    ws.cell(row=row, column=5, value="Rater A: retrieval").font = HEADER_FONT
    ws.cell(row=row, column=6, value="Rater B: gio_mode").font = HEADER_FONT
    ws.cell(row=row, column=7, value="Rater B: gn_level").font = HEADER_FONT
    ws.cell(row=row, column=8, value="Rater B: retrieval").font = HEADER_FONT
    ws.cell(row=row, column=9, value="Diskussion / Konsens").font = HEADER_FONT

    for col in range(1, 10):
        ws.cell(row=row, column=col).fill = CALIBRATION_FILL

    row += 1

    for _, prompt in calib_prompts.iterrows():
        ws.cell(row=row, column=1, value=prompt["prompt_id"])
        cell = ws.cell(row=row, column=2, value=prompt["prompt_text"])
        cell.alignment = WRAP_ALIGNMENT
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = CALIBRATION_FILL
        row += 1

    # Dropdowns fuer Kalibrierung
    n = len(calib_prompts)
    dv_mode = DataValidation(type="list",
                             formula1=f'"{",".join(config.DROPDOWN_GIO_MODES)}"',
                             allow_blank=True)
    dv_mode.sqref = f"C5:C{4 + n} F5:F{4 + n}"
    ws.add_data_validation(dv_mode)

    dv_gn = DataValidation(type="list",
                           formula1=f'"{",".join(config.DROPDOWN_GN_LEVELS)}"',
                           allow_blank=True)
    dv_gn.sqref = f"D5:D{4 + n} G5:G{4 + n}"
    ws.add_data_validation(dv_gn)

    dv_ret = DataValidation(type="list",
                            formula1=f'"{",".join(config.DROPDOWN_RETRIEVAL)}"',
                            allow_blank=True)
    dv_ret.sqref = f"E5:E{4 + n} H5:H{4 + n}"
    ws.add_data_validation(dv_ret)

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 55
    for letter in "CDEFGHI":
        ws.column_dimensions[letter].width = 18


def create_baseline_sheet(wb, study_prompts, baseline_df):
    """Erstelle das Baseline-Sheet."""
    ws = wb.create_sheet("Baseline")

    headers = ["prompt_id", "prompt_text", "baseline_prediction", "triggered_rules"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL

    if baseline_df is not None:
        for i, (_, row_data) in enumerate(baseline_df.iterrows(), start=2):
            ws.cell(row=i, column=1, value=row_data.get("prompt_id", ""))
            ws.cell(row=i, column=2, value=row_data.get("prompt_text", ""))
            ws.cell(row=i, column=3, value=row_data.get("baseline_prediction", ""))
            ws.cell(row=i, column=4, value=row_data.get("triggered_rules", ""))
    else:
        ws.cell(row=2, column=1,
                value="[Baseline-Daten noch nicht verfuegbar. Bitte AP3 zuerst ausfuehren.]")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 40


def create_metadata_sheet(wb, all_prompts):
    """Erstelle das Metadaten-Sheet (nur fuer Studienleitung)."""
    ws = wb.create_sheet("Metadaten")
    ws.sheet_properties.tabColor = META_TAB_COLOR

    # Warnung
    ws.cell(row=1, column=1,
            value="NICHT FUER RATER — Nur fuer Studienleitung").font = Font(
        bold=True, size=14, color="FF0000")
    ws.cell(row=2, column=1,
            value="Dieses Sheet enthaelt die Gold-Zuordnungen und darf waehrend "
                  "der Annotation nicht eingesehen werden.").font = Font(italic=True)

    row = 4
    headers = ["prompt_id", "source", "block", "subtype",
               "gio_mode_estimate", "justification"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    row += 1
    for _, p in all_prompts.iterrows():
        ws.cell(row=row, column=1, value=p.get("prompt_id", ""))
        ws.cell(row=row, column=2, value=p.get("source", ""))
        ws.cell(row=row, column=3, value=p.get("block", ""))
        ws.cell(row=row, column=4, value=p.get("subtype", ""))
        ws.cell(row=row, column=5, value=p.get("gio_mode_estimate", ""))
        cell = ws.cell(row=row, column=6, value=p.get("justification", ""))
        cell.alignment = WRAP_ALIGNMENT
        row += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 60


def main():
    print("=" * 60)
    print("AP4: Annotations-Spreadsheet erstellen")
    print("=" * 60)

    # Prompts laden
    print(f"\n[1/3] Lade Prompts aus {config.SAMPLED_PROMPTS_PATH}...")
    if not config.SAMPLED_PROMPTS_PATH.exists():
        print(f"FEHLER: {config.SAMPLED_PROMPTS_PATH} nicht gefunden.")
        print("Bitte zuerst AP2 abschliessen.")
        sys.exit(1)

    all_prompts = pd.read_csv(config.SAMPLED_PROMPTS_PATH)
    study_prompts = all_prompts[all_prompts["block"] != "calibration"].copy()
    calib_prompts = all_prompts[all_prompts["block"] == "calibration"].copy()

    print(f"       {len(study_prompts)} Studien-Prompts, {len(calib_prompts)} Kalibrierungs-Prompts.")

    # Baseline laden (optional)
    baseline_df = None
    if config.BASELINE_PREDICTIONS_PATH.exists():
        baseline_df = pd.read_csv(config.BASELINE_PREDICTIONS_PATH)
        print(f"       Baseline-Vorhersagen geladen ({len(baseline_df)} Eintraege).")
    else:
        print("       HINWEIS: Baseline-Daten nicht vorhanden. Sheet wird als Platzhalter erstellt.")

    # Workbook erstellen
    print("\n[2/3] Erstelle Excel-Workbook...")
    wb = Workbook()
    # Default-Sheet entfernen
    wb.remove(wb.active)

    # Sheets erstellen
    print("       Sheet 'Codebook'...")
    create_codebook_sheet(wb)

    print("       Sheet 'Kalibrierung'...")
    create_calibration_sheet(wb, calib_prompts)

    print("       Sheet 'Rater_A' (Seed=42)...")
    create_rater_sheet(wb, "Rater_A", study_prompts, seed=42)

    print("       Sheet 'Rater_B' (Seed=123)...")
    create_rater_sheet(wb, "Rater_B", study_prompts, seed=123)

    print("       Sheet 'Baseline'...")
    create_baseline_sheet(wb, study_prompts, baseline_df)

    print("       Sheet 'Metadaten'...")
    create_metadata_sheet(wb, all_prompts)

    # Speichern
    print(f"\n[3/3] Speichere nach {config.ANNOTATION_XLSX_PATH}...")
    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(config.ANNOTATION_XLSX_PATH)
    print(f"       Gespeichert: {config.ANNOTATION_XLSX_PATH}")

    # Zusammenfassung
    print("\n--- Zusammenfassung ---")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")
    print(f"  Studien-Prompts: {len(study_prompts)}")
    print(f"  Kalibrierungs-Prompts: {len(calib_prompts)}")
    print(f"  Rater_A Reihenfolge != Rater_B Reihenfolge: Ja (Seeds 42 vs 123)")
    print(f"  Dropdowns: gio_mode, i_gap, t_decay, e_spec, v_volatility, gn_level, retrieval_judgment, confidence")
    print("\nFertig. Bitte die Datei in Excel/LibreOffice oeffnen und Dropdowns testen.")


if __name__ == "__main__":
    main()
